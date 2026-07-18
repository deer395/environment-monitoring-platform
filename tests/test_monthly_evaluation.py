from pathlib import Path

import openpyxl
import pandas as pd

from src.monthly_evaluation_dataset import scan_monthly_workbooks, stable_sample_id
from src.monthly_evaluation_export import export_all
from src.variable_registry import list_enabled_variables


def _write_station(path: Path, invalid_depth: bool = False):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for key in list_enabled_variables():
            sheet = f"{key}_monthly_statistics"
            # Simulate the mandatory Excel 31-character truncation.
            sheet = sheet[:31]
            columns = {"year_month": ["2026-03", "2026-01", "2026-02"], "monthly_mean": [3.0, 1.0, 2.0], "monthly_std": [.3, .1, .2]}
            if invalid_depth and key == "depth": columns.pop("monthly_std")
            pd.DataFrame(columns).to_excel(writer, sheet_name=sheet, index=False)
        pd.DataFrame({"year_month": ["2026-01"], "daily_range": [99]}).to_excel(writer, sheet_name="depth_daily_range_statistics", index=False)


def test_scan_and_export_monthly_evaluation_materials(tmp_path):
    for name in ("station_a.xlsx", "station_b.xlsx", "station_c.xlsx"):
        _write_station(tmp_path / name)
    complete_samples, complete_warnings, complete_count = scan_monthly_workbooks(tmp_path)
    assert complete_count == 3
    assert len(complete_samples) == 27
    assert not complete_warnings
    _write_station(tmp_path / "station_c.xlsx", invalid_depth=True)
    (tmp_path / "broken.xlsx").write_text("not an Excel file", encoding="utf-8")
    samples, warnings, count = scan_monthly_workbooks(tmp_path)
    assert count == 4
    assert len(samples) == 26  # 27 expected station-variable combinations, one invalid monthly sheet.
    assert any("missing required columns: monthly_std" in warning["reason"] for warning in warnings)
    assert any("unable to read workbook" in warning["reason"] for warning in warnings)
    depth = next(sample for sample in samples if sample["station_name"] == "station_a" and sample["variable_key"] == "depth")
    assert depth["monthly"]["year_month"].tolist() == ["2026-01", "2026-02", "2026-03"]
    assert depth["monthly"]["monthly_std"].tolist() == [.1, .2, .3]
    assert depth["baseline"]["current_monthly_narrative"]
    assert stable_sample_id("station_a", "depth", "2026-01", "2026-03") == depth["sample_id"]
    outputs = export_all(samples, warnings, tmp_path / "out")
    assert outputs["figures"] == 26
    assert (tmp_path / "out" / "baseline_predictions.json").exists()
    annotation = openpyxl.load_workbook(tmp_path / "out" / "annotation_template.xlsx")
    sheet = annotation["月际标注"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.cell(2, headers.index("human_primary_pattern") + 1).value is None
    assert len(sheet.data_validations.dataValidation) >= 10
    assert "填写说明" in annotation.sheetnames
