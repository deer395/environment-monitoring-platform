"""Export blind monthly figures, baselines, manifests, and annotation workbooks."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

MANIFEST_COLUMNS = ["sample_id", "station_name", "workbook_name", "variable_key", "variable_name", "unit", "analysis_start_month", "analysis_end_month", "month_count", "monthly_mean_min", "monthly_mean_max", "monthly_mean_range", "monthly_std_mean", "monthly_std_max", "current_stage_count", "current_direction_change_count", "current_primary_pattern", "selected_for_annotation", "dataset_split", "data_status", "notes"]
ANNOTATION_COLUMNS = ["sample_id", "station_name", "variable_key", "variable_name", "unit", "analysis_start_month", "analysis_end_month", "month_count", "blind_figure_path", "human_primary_pattern", "human_trend_direction", "human_change_point_present", "human_change_point_month", "human_key_stages", "human_should_describe_short_reversal", "human_highest_month", "human_lowest_month", "human_max_std_month", "human_reference_summary", "human_annotation_confidence", "annotation_status", "dataset_split", "selected_for_annotation", "baseline_overall_rating", "baseline_error_types", "baseline_notes"]


def _json_default(value: Any):
    if hasattr(value, "item"): return value.item()
    raise TypeError(f"not JSON serializable: {type(value)!r}")


def create_blind_figures(samples: list[dict], output_dir: Path) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    for sample in samples:
        monthly = sample["monthly"]
        x = pd.to_datetime(monthly["year_month"], errors="coerce")
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.errorbar(x, monthly["monthly_mean"], yerr=monthly["monthly_std"], fmt="o-", color="#1f77b4", capsize=3, linewidth=1.5)
        ax.set_title(f"{sample['variable_name']}（{sample['unit']}）")
        ax.set_xlabel("月份"); ax.set_ylabel(sample["unit"])
        ax.grid(axis="y", alpha=.25); fig.autofmt_xdate(); fig.tight_layout()
        path = output_dir / f"{sample['sample_id']}.png"
        fig.savefig(path, dpi=300); plt.close(fig)
        sample["blind_figure_path"] = path.as_posix()
    return len(samples)


def _add_validation(sheet, column: int, options: list[str], rows: int) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{sheet.cell(2, column).coordinate}:{sheet.cell(rows + 1, column).coordinate}")


def create_annotation_template(samples: list[dict], output_path: Path) -> None:
    book = Workbook(); sheet = book.active; sheet.title = "月际标注"
    sheet.append(ANNOTATION_COLUMNS)
    auto_fields = set(ANNOTATION_COLUMNS[:9]) | {"dataset_split", "selected_for_annotation"}
    for sample in samples:
        row = [sample.get(column, "") if column in auto_fields else "" for column in ANNOTATION_COLUMNS]
        sheet.append(row)
    for cell in sheet[1]: cell.font = Font(bold=True); cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 26, "B": 16, "C": 18, "D": 14, "E": 10, "F": 14, "G": 14, "H": 11, "I": 48, "N": 46, "S": 50, "Z": 50}
    for column, width in widths.items(): sheet.column_dimensions[column].width = width
    rules = {"human_primary_pattern": ["stable", "increasing", "decreasing", "rise_then_fall", "fall_then_rise", "multi_stage", "change_point", "unclear"], "human_trend_direction": ["increase", "decrease", "stable", "mixed", "unclear"], "human_change_point_present": ["yes", "no", "unclear"], "human_should_describe_short_reversal": ["yes", "no", "unclear"], "human_annotation_confidence": ["high", "medium", "low"], "annotation_status": ["not_started", "in_progress", "completed", "needs_review"], "dataset_split": ["development", "holdout"], "selected_for_annotation": ["yes", "no"], "baseline_overall_rating": ["correct", "mostly_correct", "minor_revision", "major_revision", "incorrect"], "baseline_error_types": ["trend_direction_error", "change_point_error", "stage_boundary_error", "over_fragmented", "over_simplified", "missing_key_stage", "unsupported_wording", "numeric_error", "other"]}
    for field, options in rules.items(): _add_validation(sheet, ANNOTATION_COLUMNS.index(field) + 1, options, len(samples))
    guide = book.create_sheet("填写说明")
    instructions = ["月际报告解读质量评估：人工标注说明", "1. 第一步只看 blind 图，不要先看 baseline_review.xlsx。", "2. 根据月平均和月标准差独立填写所有 human_* 字段。", "3. 完成人工标签后，再查看当前算法结果并填写 baseline 评价字段。", "4. 人工文字不要求与程序逐字一致；重点比较方向、阶段、转折和关键事实。", "5. 月标准差只表示月内波动，不表示月平均趋势。", "6. 不对环境成因进行解释。", "7. development 可用于后续调算法；holdout 在算法修改期间不用于调参。", "8. human_key_stages 示例：2025-08~2025-10: decrease; 2025-10~2026-02: stable; 2026-02~2026-06: increase", "9. baseline_error_types 多项时用分号分隔。"]
    for line in instructions: guide.append([line])
    guide.column_dimensions["A"].width = 110; guide["A1"].font = Font(bold=True)
    output_path.parent.mkdir(parents=True, exist_ok=True); book.save(output_path)


def export_all(samples: list[dict], warnings: list[dict[str, str]], output_dir: str | Path) -> dict[str, Path | int]:
    output = Path(output_dir); blind = output / "figures" / "blind"; output.mkdir(parents=True, exist_ok=True); (output / "results").mkdir(exist_ok=True)
    figure_count = create_blind_figures(samples, blind)
    manifest = output / "sample_manifest.csv"
    pd.DataFrame([{key: sample.get(key, "") for key in MANIFEST_COLUMNS} for sample in samples], columns=MANIFEST_COLUMNS).to_csv(manifest, index=False, encoding="utf-8-sig")
    predictions = []
    for sample in samples:
        predictions.append({"sample_id": sample["sample_id"], "station_name": sample["station_name"], "variable_key": sample["variable_key"], "months": sample["monthly"]["year_month"].tolist(), "monthly_means": sample["monthly"]["monthly_mean"].tolist(), "monthly_stds": sample["monthly"]["monthly_std"].tolist(), **sample["baseline"], "algorithm_version": "v4.2-frozen"})
    baseline_path = output / "baseline_predictions.json"; baseline_path.write_text(json.dumps(predictions, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")
    review_columns = ["sample_id", "station_name", "variable_name", "current_primary_pattern", "current_stages", "current_highest_month", "current_lowest_month", "current_max_std_month", "current_monthly_narrative", "blind_figure_path"]
    review_rows = [{column: (json.dumps(sample["baseline"].get(column), ensure_ascii=False) if column == "current_stages" else sample.get(column, sample["baseline"].get(column, ""))) for column in review_columns} for sample in samples]
    pd.DataFrame(review_rows, columns=review_columns).to_excel(output / "baseline_review.xlsx", index=False)
    create_annotation_template(samples, output / "annotation_template.xlsx")
    return {"manifest": manifest, "baseline": baseline_path, "annotation": output / "annotation_template.xlsx", "review": output / "baseline_review.xlsx", "figures": figure_count, "warnings": len(warnings)}
